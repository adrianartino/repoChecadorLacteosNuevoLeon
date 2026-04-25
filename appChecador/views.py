import os
from datetime import datetime, timedelta
from io import BytesIO

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth import authenticate, login as auth_login
from django.contrib.auth import logout as auth_logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Count, Q, Sum
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import Departamento, Empleado, Horario, JornadaAsistencia, RegistroAsistencia
from .attendance_engine import reprocesar_jornadas_empleado
from .sync_hikvision import sincronizar_checadas


LOGO_REPORTE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'static',
    'images',
    'fondoLogin',
    'logoLacteos.jpeg',
)


def _dibujar_marco_pdf(canvas, doc):
    canvas.saveState()
    borde = 0.22 * inch
    ancho, alto = doc.pagesize
    canvas.setStrokeColor(colors.HexColor('#CBD5E1'))
    canvas.setLineWidth(0.8)
    canvas.roundRect(
        borde,
        borde,
        ancho - (borde * 2),
        alto - (borde * 2),
        10,
        stroke=1,
        fill=0,
    )
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.HexColor('#64748B'))
    canvas.drawRightString(ancho - borde, borde - 0.04 * inch, f'Pagina {canvas.getPageNumber()}')
    canvas.restoreState()


def _formatear_hora_local(dt):
    """Hora HH:MM en zona del proyecto; acepta naive (p. ej. combine IMSS) o aware."""
    if dt is None:
        return '--:--'
    if not isinstance(dt, datetime):
        return '--:--'
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return timezone.localtime(dt).strftime('%H:%M')


def _formatear_fecha_hora_excel_jornada(dt, vacio):
    """Fecha/hora para celdas de Excel de reportes de jornadas (incl. IMSS con combine naive)."""
    if dt is None:
        return vacio
    if not isinstance(dt, datetime):
        return vacio
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return timezone.localtime(dt).strftime('%Y-%m-%d %I:%M:%S %p')


def _formatear_hora_am_pm_jornada(dt, vacio):
    if dt is None:
        return vacio
    if not isinstance(dt, datetime):
        return vacio
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return timezone.localtime(dt).strftime('%I:%M %p')


def login(request):
    # 1. Verificamos si el usuario ya tiene una sesión iniciada
    if 'usuario_id' in request.session:
        return redirect('inicio')

    if request.method == 'POST':
        usuario = request.POST.get('username')
        contra = request.POST.get('password')
        
        user = authenticate(request, username=usuario, password=contra)
        
        if user is not None:
            # 2. auth_login crea la sesión automáticamente en la base de datos y el navegador
            auth_login(request, user) 
            
            # --- CREAMOS VARIABLES GLOBALES DE SESIÓN ---
            request.session['usuario_id'] = user.id
            request.session['usuario_nombre'] = user.username
            request.session['es_admin'] = user.is_superuser
            # --------------------------------------------
            
            return redirect('inicio') 
        else:
            print("malo")
            return render(request, 'appChecador/login/login.html', {'error': True})
            
    return render(request, 'appChecador/login/login.html')

def logout(request):
    # Borra las variables de sesión personalizadas y la cookie de Django
    request.session.flush() 
    auth_logout(request)
    # Redirige al login
    return redirect('login')

def inicio(request):
    if 'usuario_id' not in request.session:
        return redirect('login')
    
    nombreUsuarioLogueado = request.session['usuario_nombre']
    hoy = timezone.localdate()

    jornadas_hoy = JornadaAsistencia.objects.filter(fecha_administrativa=hoy).select_related(
        'empleado', 'empleado__departamento', 'horario'
    )
    registros_hoy = RegistroAsistencia.objects.filter(fecha_administrativa=hoy).select_related('empleado')

    total_empleados = Empleado.objects.count()
    empleados_activos = Empleado.objects.filter(activo=True).count()
    empleados_sin_configurar = Empleado.objects.filter(
        Q(departamento__isnull=True) | Q(horario__isnull=True)
    ).count()

    total_jornadas_hoy = jornadas_hoy.count()
    jornadas_completas_hoy = jornadas_hoy.filter(salida_real__isnull=False).count()
    jornadas_sin_salida_hoy = jornadas_hoy.filter(entrada_real__isnull=False, salida_real__isnull=True).count()
    retardos_hoy = jornadas_hoy.filter(retardo=True).count()
    horas_extra_hoy = jornadas_hoy.aggregate(total=Sum('horas_extra'))['total'] or 0
    checadas_hoy = registros_hoy.count()
    duplicadas_hoy = registros_hoy.filter(observaciones__icontains='Duplicada por ventana').count()

    ultima_checada = registros_hoy.order_by('-fecha_hora_real').first() or RegistroAsistencia.objects.order_by('-fecha_hora_real').first()

    resumen_departamentos = list(
        jornadas_hoy.values('empleado__departamento__nombre')
        .annotate(
            asistencias=Count('id'),
            retardos=Count('id', filter=Q(retardo=True)),
            horas_extra=Sum('horas_extra'),
        )
        .order_by('-asistencias', 'empleado__departamento__nombre')[:8]
    )

    for item in resumen_departamentos:
        item['departamento'] = item['empleado__departamento__nombre'] or 'Sin departamento'
        item['horas_extra'] = item['horas_extra'] or 0

    incidencias = []
    for jornada in jornadas_hoy.filter(Q(salida_real__isnull=True) | Q(retardo=True)).order_by('empleado__nombre', 'entrada_real')[:8]:
        if jornada.salida_real is None:
            tipo = 'Sin salida'
            detalle = 'La jornada sigue abierta.'
        elif jornada.retardo:
            tipo = 'Retardo'
            detalle = f'{jornada.minutos_retardo} min de retardo.'
        else:
            tipo = 'Revision'
            detalle = 'Incidencia operativa.'
        incidencias.append({
            'empleado': jornada.empleado.nombre,
            'departamento': jornada.empleado.departamento.nombre if jornada.empleado.departamento else 'Sin departamento',
            'tipo': tipo,
            'detalle': detalle,
            'fecha': jornada.fecha_administrativa,
        })

    actividad_reciente = list(
        registros_hoy.order_by('-fecha_hora_real')[:8]
    )

    return render(
        request,
        'appChecador/inicio/inicio.html',
        {
            'nombre_usuario': nombreUsuarioLogueado,
            'fecha_hoy': hoy,
            'total_empleados': total_empleados,
            'empleados_activos': empleados_activos,
            'empleados_sin_configurar': empleados_sin_configurar,
            'total_jornadas_hoy': total_jornadas_hoy,
            'jornadas_completas_hoy': jornadas_completas_hoy,
            'jornadas_sin_salida_hoy': jornadas_sin_salida_hoy,
            'retardos_hoy': retardos_hoy,
            'horas_extra_hoy': horas_extra_hoy,
            'checadas_hoy': checadas_hoy,
            'duplicadas_hoy': duplicadas_hoy,
            'ultima_checada': ultima_checada,
            'resumen_departamentos': resumen_departamentos,
            'incidencias': incidencias,
            'actividad_reciente': actividad_reciente,
        }
    )


def _parsear_fecha(valor, predeterminada):
    if not valor:
        return predeterminada

    try:
        return timezone.datetime.strptime(valor, '%Y-%m-%d').date()
    except ValueError:
        return predeterminada


def _construir_filtros_asistencias(request):
    hoy = timezone.localdate()
    datos = request.POST if request.method == 'POST' else request.GET
    fecha = _parsear_fecha(datos.get('fecha'), hoy)
    departamento_id = datos.get('departamento') or ''
    empleado_id = datos.get('empleado') or ''
    horario_id = datos.get('horario') or ''

    filtros = {
        'fecha': fecha,
        'departamento': departamento_id,
        'empleado': empleado_id,
        'horario': horario_id,
    }

    jornadas = (
        JornadaAsistencia.objects.filter(fecha_administrativa=fecha)
        .select_related('empleado', 'empleado__departamento', 'horario')
        .order_by('empleado__nombre', 'entrada_real')
    )

    if departamento_id:
        jornadas = jornadas.filter(empleado__departamento_id=departamento_id)
    if empleado_id:
        jornadas = jornadas.filter(empleado_id=empleado_id)
    if horario_id:
        jornadas = jornadas.filter(horario_id=horario_id)

    return filtros, jornadas


def _obtener_catalogos_asistencias():
    return {
        'departamentos': Departamento.objects.order_by('nombre'),
        'empleados': Empleado.objects.order_by('nombre'),
        'horarios': Horario.objects.order_by('nombre_turno'),
    }


def _agrupar_horarios_por_departamento():
    grupos = {}
    for horario in Horario.objects.order_by('nombre_turno'):
        nombre = horario.nombre_turno
        departamento = nombre.split(' - ')[0] if ' - ' in nombre else 'Otros'
        grupos.setdefault(departamento, []).append(horario)
    return sorted(grupos.items(), key=lambda item: item[0])


def _construir_filtros_resumen(request):
    hoy = timezone.localdate()
    datos = request.POST if request.method == 'POST' else request.GET
    fecha = _parsear_fecha(datos.get('fecha'), hoy)
    fecha_base_semana = _parsear_fecha(datos.get('fecha_base'), hoy)
    departamento_id = datos.get('departamento') or ''
    empleado_id = datos.get('empleado') or ''

    return {
        'fecha': fecha,
        'fecha_base': fecha_base_semana,
        'departamento': departamento_id,
        'empleado': empleado_id,
    }


def _aplicar_filtros_resumen(jornadas, filtros):
    if filtros['departamento']:
        jornadas = jornadas.filter(empleado__departamento_id=filtros['departamento'])
    if filtros['empleado']:
        jornadas = jornadas.filter(empleado_id=filtros['empleado'])
    return jornadas


def _resumen_departamentos(jornadas):
    resumen = {}
    for jornada in jornadas:
        nombre = jornada.empleado.departamento.nombre if jornada.empleado.departamento else 'Sin departamento'
        item = resumen.setdefault(nombre, {
            'departamento': nombre,
            'total_jornadas': 0,
            'completas': 0,
            'retardos': 0,
            'con_extra': 0,
            'sin_salida': 0,
        })
        item['total_jornadas'] += 1
        if jornada.entrada_real and jornada.salida_real:
            item['completas'] += 1
        if jornada.retardo:
            item['retardos'] += 1
        if jornada.horas_extra > 0:
            item['con_extra'] += 1
        if jornada.entrada_real and not jornada.salida_real:
            item['sin_salida'] += 1
    return sorted(resumen.values(), key=lambda item: item['departamento'])


def _resumen_empleados(jornadas):
    resumen = {}
    for jornada in jornadas:
        item = resumen.setdefault(jornada.empleado_id, {
            'empleado_id': jornada.empleado_id,
            'id_biometrico': jornada.empleado.id_biometrico,
            'empleado': jornada.empleado.nombre,
            'departamento': jornada.empleado.departamento.nombre if jornada.empleado.departamento else 'Sin departamento',
            'horario': jornada.horario.nombre_turno if jornada.horario else 'Sin horario',
            'entrada': jornada.entrada_real,
            'salida': jornada.salida_real,
            'retardo': jornada.retardo,
            'horas_extra': jornada.horas_extra,
            'minutos_extra': jornada.minutos_extra,
        })
        if jornada.horas_extra > item['horas_extra']:
            item['horas_extra'] = jornada.horas_extra
            item['minutos_extra'] = jornada.minutos_extra
        if jornada.retardo:
            item['retardo'] = True
        if jornada.entrada_real and (item['entrada'] is None or jornada.entrada_real < item['entrada']):
            item['entrada'] = jornada.entrada_real
        if jornada.salida_real and (item['salida'] is None or jornada.salida_real > item['salida']):
            item['salida'] = jornada.salida_real
        if jornada.horario and item['horario'] == 'Sin horario':
            item['horario'] = jornada.horario.nombre_turno
    return sorted(resumen.values(), key=lambda item: item['empleado'])


def _resumen_semana_operativa(jornadas, modo_imss=False):
    resumen = {}
    for jornada in jornadas:
        item = resumen.setdefault(jornada.empleado_id, {
            'empleado': jornada.empleado.nombre,
            'id_biometrico': jornada.empleado.id_biometrico,
            'departamento': jornada.empleado.departamento.nombre if jornada.empleado.departamento else 'Sin departamento',
            'total_jornadas': 0,
            'completas': 0,
            'retardos': 0,
            'horas_extra': 0,
            'minutos_extra': 0,
            'dias_con_extra': 0,
            'sin_salida': 0,
        })
        item['total_jornadas'] += 1
        if jornada.entrada_real and jornada.salida_real:
            item['completas'] += 1
        if jornada.entrada_real and not jornada.salida_real:
            item['sin_salida'] += 1
        if not modo_imss and jornada.retardo:
            item['retardos'] += 1
        if not modo_imss and jornada.horas_extra > 0:
            item['horas_extra'] += jornada.horas_extra
            item['minutos_extra'] += jornada.minutos_extra
            item['dias_con_extra'] += 1
    return sorted(resumen.values(), key=lambda item: item['empleado'])


def _detalle_semana_operativa(jornadas, inicio_semana, fin_semana, modo_imss=False):
    dias_semana = []
    fecha_cursor = inicio_semana
    abreviaturas = ['J', 'V', 'S', 'D', 'L', 'M', 'Mi']
    indice = 0
    while fecha_cursor <= fin_semana:
        dias_semana.append({
            'fecha': fecha_cursor,
            'abreviatura': abreviaturas[indice],
        })
        fecha_cursor += timedelta(days=1)
        indice += 1

    empleados = {}
    for jornada in jornadas:
        item = empleados.setdefault(jornada.empleado_id, {
            'empleado': jornada.empleado.nombre,
            'id_biometrico': jornada.empleado.id_biometrico,
            'departamento': jornada.empleado.departamento.nombre if jornada.empleado.departamento else 'Sin departamento',
            'dias': {},
            'dias_trabajados': 0,
            'horas_extra_semana': 0,
        })
        entrada = _valor_reporte(jornada, 'entrada_real', modo_imss=modo_imss)
        salida = _valor_reporte(jornada, 'salida_real', modo_imss=modo_imss)
        horas_extra = _valor_reporte(jornada, 'horas_extra', modo_imss=modo_imss)
        minutos_extra = _valor_reporte(jornada, 'minutos_extra', modo_imss=modo_imss)
        item['dias'][jornada.fecha_administrativa] = {
            'entrada': entrada,
            'salida': salida,
            'horas_extra': horas_extra,
            'minutos_extra': minutos_extra,
        }
        item['dias_trabajados'] += 1
        item['horas_extra_semana'] += horas_extra

    filas = []
    for empleado_id, item in empleados.items():
        columnas = []
        for dia in dias_semana:
            registro = item['dias'].get(dia['fecha'])
            if not registro:
                columnas.append({
                    'texto': 'D',
                    'entrada': None,
                    'salida': None,
                    'horas_extra': 0,
                    'minutos_extra': 0,
                })
                continue

            columnas.append({
                'texto': '',
                'entrada': registro['entrada'],
                'salida': registro['salida'],
                'horas_extra': registro['horas_extra'],
                'minutos_extra': registro['minutos_extra'],
            })

        filas.append({
            'empleado': item['empleado'],
            'id_biometrico': item['id_biometrico'],
            'departamento': item['departamento'],
            'columnas': columnas,
            'dias_trabajados': item['dias_trabajados'],
            'horas_extra_semana': item['horas_extra_semana'],
        })

    return dias_semana, sorted(filas, key=lambda fila: fila['empleado'])


def _texto_filtros_semana(filtros, inicio_semana, fin_semana):
    departamento = 'Todos'
    empleado = 'Todos'

    if filtros.get('departamento'):
        departamento_obj = Departamento.objects.filter(id=filtros['departamento']).first()
        if departamento_obj:
            departamento = departamento_obj.nombre

    if filtros.get('empleado'):
        empleado_obj = Empleado.objects.filter(id=filtros['empleado']).first()
        if empleado_obj:
            empleado = empleado_obj.nombre

    return (
        f"Fecha base: {filtros['fecha_base'].isoformat()} | "
        f"Periodo: {inicio_semana.isoformat()} al {fin_semana.isoformat()} | "
        f"Departamento: {departamento} | "
        f"Empleado: {empleado}"
    )


def _estado_jornada(jornada):
    if jornada.entrada_real and jornada.salida_real:
        if jornada.retardo:
            return 'Completa con retardo'
        if jornada.horas_extra:
            return 'Completa con extra'
        return 'Completa'
    if jornada.entrada_real and not jornada.salida_real:
        return 'Sin salida'
    if jornada.salida_real and not jornada.entrada_real:
        return 'Sin entrada'
    return 'Sin checadas'


def _nombre_archivo_reporte(prefijo, fecha):
    return f'{prefijo}_{fecha.isoformat()}.xlsx'


def _entrada_imss(jornada):
    if jornada.horario:
        naive = timezone.datetime.combine(jornada.fecha_administrativa, jornada.horario.hora_entrada)
        return timezone.make_aware(naive, timezone.get_current_timezone())
    return jornada.entrada_real


def _salida_imss(jornada):
    if not jornada.horario:
        return jornada.salida_real

    fecha_salida = jornada.fecha_administrativa
    if jornada.horario.hora_entrada > jornada.horario.hora_salida:
        fecha_salida = jornada.fecha_administrativa + timedelta(days=1)
    naive = timezone.datetime.combine(fecha_salida, jornada.horario.hora_salida)
    return timezone.make_aware(naive, timezone.get_current_timezone())


def _valor_reporte(jornada, campo, modo_imss=False):
    if not modo_imss:
        return getattr(jornada, campo)

    if campo == 'entrada_real':
        return _entrada_imss(jornada)
    if campo == 'salida_real':
        return _salida_imss(jornada)
    if campo == 'retardo':
        return False
    if campo == 'minutos_extra':
        return 0
    if campo == 'horas_extra':
        return 0
    return getattr(jornada, campo)


def _crear_excel_reporte_jornadas(jornadas, filtros, nombre_reporte, prefijo_archivo, modo_imss=False):
    workbook = Workbook()
    hoja = workbook.active
    hoja.title = nombre_reporte[:31]

    hoja.merge_cells('A1:H1')
    hoja['A1'] = 'LACTEOS NUEVO LEON'
    hoja['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    hoja['A1'].fill = PatternFill('solid', fgColor='111827')

    hoja.merge_cells('A2:H2')
    sufijo = ' | Formato IMSS' if modo_imss else ''
    hoja['A2'] = f"Reporte: {nombre_reporte}{sufijo} | Fecha administrativa: {filtros['fecha'].isoformat()}"
    hoja['A2'].font = Font(bold=True, size=12)

    encabezados = [
        'ID Biometrico',
        'Empleado',
        'Departamento',
        'Horario',
        'Entrada',
        'Salida',
        'Retardo',
        'Horas Extra',
    ]
    hoja.append([])
    hoja.append(encabezados)

    for celda in hoja[4]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor='1D4ED8')

    for jornada in jornadas:
        entrada = _valor_reporte(jornada, 'entrada_real', modo_imss=modo_imss)
        salida = _valor_reporte(jornada, 'salida_real', modo_imss=modo_imss)
        hoja.append([
            jornada.empleado.id_biometrico,
            jornada.empleado.nombre,
            jornada.empleado.departamento.nombre if jornada.empleado.departamento else 'Sin departamento',
            jornada.horario.nombre_turno if jornada.horario else 'Sin horario',
            _formatear_fecha_hora_excel_jornada(entrada, 'Sin entrada'),
            _formatear_fecha_hora_excel_jornada(salida, 'Sin salida'),
            'Si' if _valor_reporte(jornada, 'retardo', modo_imss=modo_imss) else 'No',
            _valor_reporte(jornada, 'horas_extra', modo_imss=modo_imss),
        ])

    for indice_columna, columna in enumerate(hoja.iter_cols(min_row=4, max_row=hoja.max_row), start=1):
        max_len = 0
        letra_columna = get_column_letter(indice_columna)
        for celda in columna:
            valor = str(celda.value or '')
            if len(valor) > max_len:
                max_len = len(valor)
        hoja.column_dimensions[letra_columna].width = min(max_len + 3, 32)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = prefijo_archivo if not modo_imss else f'{prefijo_archivo}_imss'
    response['Content-Disposition'] = f'attachment; filename="{_nombre_archivo_reporte(nombre_archivo, filtros["fecha"])}"'
    workbook.save(response)
    return response


def _crear_pdf_reporte_jornadas(jornadas, filtros, nombre_reporte, prefijo_archivo, modo_imss=False):
    buffer = BytesIO()
    response = HttpResponse(content_type='application/pdf')
    nombre_archivo = prefijo_archivo if not modo_imss else f'{prefijo_archivo}_imss'
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}_{filtros["fecha"].isoformat()}.pdf"'
    documento = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=0.50 * inch,
        rightMargin=0.50 * inch,
        topMargin=0.50 * inch,
        bottomMargin=0.50 * inch,
    )

    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        'TituloReporte',
        parent=estilos['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#111827'),
        spaceAfter=4,
    )
    subtitulo = ParagraphStyle(
        'SubtituloReporte',
        parent=estilos['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#4B5563'),
        leading=14,
    )

    elementos = []
    encabezado = []
    if os.path.exists(LOGO_REPORTE):
        encabezado.append(Image(LOGO_REPORTE, width=1.0 * inch, height=1.0 * inch))
    else:
        encabezado.append(Spacer(1, 1.0 * inch))

    texto_encabezado = [
        Paragraph('LACTEOS NUEVO LEON S.P.P DE R.L.', titulo),
        Paragraph(
            f'{nombre_reporte}{" - Formato IMSS" if modo_imss else ""}',
            ParagraphStyle(
            'NombreReporte',
            parent=estilos['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=13,
            textColor=colors.HexColor('#1D4ED8'),
            spaceAfter=6,
        )),
        Paragraph(
            f"Fecha administrativa: {filtros['fecha'].isoformat()}<br/>"
            f"Generado: {timezone.localtime().strftime('%Y-%m-%d %I:%M:%S %p')}",
            subtitulo,
        ),
    ]
    encabezado.append(texto_encabezado)

    tabla_encabezado = Table([encabezado], colWidths=[1.1 * inch, 8.4 * inch])
    tabla_encabezado.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elementos.append(tabla_encabezado)
    elementos.append(Spacer(1, 0.25 * inch))

    datos = [[
        'ID',
        'Empleado',
        'Departamento',
        'Horario',
        'Entrada',
        'Salida',
        'Estado',
        'Extra',
    ]]

    for jornada in jornadas:
        entrada = _valor_reporte(jornada, 'entrada_real', modo_imss=modo_imss)
        salida = _valor_reporte(jornada, 'salida_real', modo_imss=modo_imss)
        datos.append([
            jornada.empleado.id_biometrico,
            jornada.empleado.nombre,
            jornada.empleado.departamento.nombre if jornada.empleado.departamento else 'Sin departamento',
            jornada.horario.nombre_turno if jornada.horario else 'Sin horario',
            _formatear_hora_am_pm_jornada(entrada, 'Sin entrada'),
            _formatear_hora_am_pm_jornada(salida, 'Sin salida'),
            'Completa' if modo_imss else _estado_jornada(jornada),
            str(_valor_reporte(jornada, 'horas_extra', modo_imss=modo_imss)),
        ])

    tabla = Table(
        datos,
        repeatRows=1,
        colWidths=[0.65 * inch, 1.60 * inch, 1.35 * inch, 1.45 * inch, 0.95 * inch, 0.95 * inch, 1.35 * inch, 0.55 * inch],
    )
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F8FAFC'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabla)

    documento.build(elementos, onFirstPage=_dibujar_marco_pdf, onLaterPages=_dibujar_marco_pdf)
    response.write(buffer.getvalue())
    buffer.close()
    return response


def _crear_excel_resumen_semana(resumen, inicio_semana, fin_semana, prefijo_archivo, modo_imss=False):
    workbook = Workbook()
    hoja = workbook.active
    hoja.title = 'Semana operativa'

    hoja.merge_cells('A1:H1')
    hoja['A1'] = 'LACTEOS NUEVO LEON'
    hoja['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    hoja['A1'].fill = PatternFill('solid', fgColor='111827')

    hoja.merge_cells('A2:H2')
    sufijo = ' | Formato IMSS' if modo_imss else ''
    hoja['A2'] = f"Reporte: Semana operativa{sufijo} | {inicio_semana.isoformat()} al {fin_semana.isoformat()}"
    hoja['A2'].font = Font(bold=True, size=12)

    hoja.append([])
    hoja.append([
        'ID Biometrico',
        'Empleado',
        'Departamento',
        'Jornadas',
        'Completas',
        'Retardos',
        'Horas Extra',
        'Dias con Extra',
    ])

    for celda in hoja[4]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor='1D4ED8')

    for fila in resumen:
        hoja.append([
            fila['id_biometrico'],
            fila['empleado'],
            fila['departamento'],
            fila['total_jornadas'],
            fila['completas'],
            0 if modo_imss else fila['retardos'],
            0 if modo_imss else fila['horas_extra'],
            0 if modo_imss else fila['dias_con_extra'],
        ])

    for indice_columna, columna in enumerate(hoja.iter_cols(min_row=4, max_row=hoja.max_row), start=1):
        max_len = 0
        letra_columna = get_column_letter(indice_columna)
        for celda in columna:
            valor = str(celda.value or '')
            if len(valor) > max_len:
                max_len = len(valor)
        hoja.column_dimensions[letra_columna].width = min(max_len + 3, 28)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = prefijo_archivo if not modo_imss else f'{prefijo_archivo}_imss'
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}_{inicio_semana.isoformat()}_{fin_semana.isoformat()}.xlsx"'
    workbook.save(response)
    return response


def _crear_pdf_resumen_semana(resumen, inicio_semana, fin_semana, prefijo_archivo, modo_imss=False):
    buffer = BytesIO()
    response = HttpResponse(content_type='application/pdf')
    nombre_archivo = prefijo_archivo if not modo_imss else f'{prefijo_archivo}_imss'
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}_{inicio_semana.isoformat()}_{fin_semana.isoformat()}.pdf"'
    documento = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=0.50 * inch,
        rightMargin=0.50 * inch,
        topMargin=0.50 * inch,
        bottomMargin=0.50 * inch,
    )

    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        'TituloSemana',
        parent=estilos['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#111827'),
        spaceAfter=4,
    )
    subtitulo = ParagraphStyle(
        'SubtituloSemana',
        parent=estilos['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#4B5563'),
        leading=14,
    )

    elementos = []
    encabezado = []
    if os.path.exists(LOGO_REPORTE):
        encabezado.append(Image(LOGO_REPORTE, width=1.0 * inch, height=1.0 * inch))
    else:
        encabezado.append(Spacer(1, 1.0 * inch))

    encabezado.append([
        Paragraph('LACTEOS NUEVO LEON S.P.P DE R.L.', titulo),
        Paragraph(
            f"Semana operativa{' - Formato IMSS' if modo_imss else ''}",
            ParagraphStyle(
                'NombreReporteSemana',
                parent=estilos['Heading2'],
                fontName='Helvetica-Bold',
                fontSize=13,
                textColor=colors.HexColor('#1D4ED8'),
                spaceAfter=6,
            )
        ),
        Paragraph(
            f"Periodo: {inicio_semana.isoformat()} al {fin_semana.isoformat()}<br/>"
            f"Generado: {timezone.localtime().strftime('%Y-%m-%d %I:%M:%S %p')}",
            subtitulo,
        ),
    ])
    tabla_encabezado = Table([encabezado], colWidths=[1.1 * inch, 8.4 * inch])
    tabla_encabezado.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elementos.append(tabla_encabezado)
    elementos.append(Spacer(1, 0.25 * inch))

    datos = [[
        'ID',
        'Empleado',
        'Departamento',
        'Jornadas',
        'Completas',
        'Retardos',
        'Horas Extra',
        'Dias Extra',
    ]]
    for fila in resumen:
        datos.append([
            fila['id_biometrico'],
            fila['empleado'],
            fila['departamento'],
            str(fila['total_jornadas']),
            str(fila['completas']),
            str(0 if modo_imss else fila['retardos']),
            str(0 if modo_imss else fila['horas_extra']),
            str(0 if modo_imss else fila['dias_con_extra']),
        ])

    tabla = Table(
        datos,
        repeatRows=1,
        colWidths=[0.8 * inch, 2.2 * inch, 1.8 * inch, 0.9 * inch, 0.9 * inch, 0.9 * inch, 1.0 * inch, 0.9 * inch],
    )
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F8FAFC'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabla)

    documento.build(elementos, onFirstPage=_dibujar_marco_pdf, onLaterPages=_dibujar_marco_pdf)
    response.write(buffer.getvalue())
    buffer.close()
    return response


def _crear_excel_semana_operativa_detallada(dias_semana, filas_semana, inicio_semana, fin_semana, prefijo_archivo, filtros, modo_imss=False):
    workbook = Workbook()
    hoja = workbook.active
    hoja.title = 'Semana operativa'

    hoja.merge_cells('A1:J1')
    hoja['A1'] = 'LACTEOS NUEVO LEON S.P.P DE R.L.'
    hoja['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    hoja['A1'].fill = PatternFill('solid', fgColor='111827')
    hoja['A1'].alignment = Alignment(horizontal='center', vertical='center')

    hoja.merge_cells('A2:J2')
    sufijo = ' | Formato IMSS' if modo_imss else ''
    hoja['A2'] = f"Semana operativa{sufijo} | {inicio_semana.isoformat()} al {fin_semana.isoformat()}"
    hoja['A2'].font = Font(bold=True, size=12)
    hoja['A2'].alignment = Alignment(horizontal='center', vertical='center')
    hoja.merge_cells('A3:J3')
    hoja['A3'] = _texto_filtros_semana(filtros, inicio_semana, fin_semana)
    hoja['A3'].font = Font(size=10)
    hoja['A3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    encabezados = ['Num', 'Nombre']
    encabezados.extend([
        f"{dia['abreviatura']}\n{dia['fecha'].strftime('%d/%m/%Y')}" for dia in dias_semana
    ])
    encabezados.append('Resumen')

    hoja.append(encabezados)

    for celda in hoja[4]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor='1D4ED8')
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    hoja.row_dimensions[4].height = 32

    for fila in filas_semana:
        valores = [fila['id_biometrico'], fila['empleado']]
        for columna in fila['columnas']:
            if columna['texto'] == 'D':
                valores.append('D')
            else:
                entrada = _formatear_hora_local(columna['entrada'])
                salida = _formatear_hora_local(columna['salida'])
                extra = f" | HE:{columna['horas_extra']}"
                if columna['horas_extra'] or columna['minutos_extra']:
                    extra = f" | HE:{columna['horas_extra']} ({columna['minutos_extra']} min)"
                valores.append(f'{entrada}/{salida}{extra}')
        valores.append(f"Dias: {fila['dias_trabajados']} | HE: {fila['horas_extra_semana']}")
        hoja.append(valores)

    for fila in hoja.iter_rows(min_row=5, max_row=hoja.max_row):
        for celda in fila:
            celda.alignment = Alignment(vertical='center', wrap_text=True)

    for indice_columna, columna in enumerate(hoja.iter_cols(min_row=4, max_row=hoja.max_row), start=1):
        max_len = 0
        letra_columna = get_column_letter(indice_columna)
        for celda in columna:
            valor = str(celda.value or '')
            if len(valor) > max_len:
                max_len = len(valor)
        hoja.column_dimensions[letra_columna].width = min(max_len + 3, 26)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = prefijo_archivo if not modo_imss else f'{prefijo_archivo}_imss'
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}_{inicio_semana.isoformat()}_{fin_semana.isoformat()}.xlsx"'
    workbook.save(response)
    return response


def _crear_pdf_semana_operativa_detallada(dias_semana, filas_semana, inicio_semana, fin_semana, prefijo_archivo, filtros, modo_imss=False):
    buffer = BytesIO()
    response = HttpResponse(content_type='application/pdf')
    nombre_archivo = prefijo_archivo if not modo_imss else f'{prefijo_archivo}_imss'
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}_{inicio_semana.isoformat()}_{fin_semana.isoformat()}.pdf"'
    documento = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=0.45 * inch,
        rightMargin=0.45 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch,
    )

    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        'TituloSemanaDetallada',
        parent=estilos['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#111827'),
        spaceAfter=4,
    )
    subtitulo = ParagraphStyle(
        'SubtituloSemanaDetallada',
        parent=estilos['BodyText'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#4B5563'),
        leading=12,
    )

    elementos = []
    encabezado = []
    if os.path.exists(LOGO_REPORTE):
        encabezado.append(Image(LOGO_REPORTE, width=0.9 * inch, height=0.9 * inch))
    else:
        encabezado.append(Spacer(1, 0.9 * inch))

    encabezado.append([
        Paragraph('LACTEOS NUEVO LEON S.P.P DE R.L.', titulo),
        Paragraph(f'Semana Operativa{" - Formato IMSS" if modo_imss else ""}', ParagraphStyle(
            'NombreReporteSemanaDetallada',
            parent=estilos['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=12,
            textColor=colors.HexColor('#1D4ED8'),
            spaceAfter=4,
        )),
        Paragraph(
            f"Periodo: {inicio_semana.isoformat()} al {fin_semana.isoformat()}<br/>"
            f"Formato semanal de asistencia y horas extra<br/>"
            f"{_texto_filtros_semana(filtros, inicio_semana, fin_semana)}",
            subtitulo,
        ),
    ])

    tabla_encabezado = Table([encabezado], colWidths=[1.0 * inch, 8.5 * inch])
    tabla_encabezado.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elementos.append(tabla_encabezado)
    elementos.append(Spacer(1, 0.18 * inch))

    datos = [['NUM', 'NOMBRE']]
    datos[0].extend(
        f"{dia['abreviatura']}\n{dia['fecha'].strftime('%d/%m/%Y')}" for dia in dias_semana
    )
    datos[0].append('RESUMEN')

    for fila in filas_semana:
        renglon = [fila['id_biometrico'], fila['empleado']]
        for columna in fila['columnas']:
            if columna['texto'] == 'D':
                renglon.append('D')
            else:
                entrada = _formatear_hora_local(columna['entrada'])
                salida = _formatear_hora_local(columna['salida'])
                extra = f"\nHE: {columna['horas_extra']}"
                if columna['horas_extra'] or columna['minutos_extra']:
                    extra = f"\nHE: {columna['horas_extra']} ({columna['minutos_extra']}m)"
                renglon.append(f'{entrada}\n{salida}{extra}')
        renglon.append(f'Dias: {fila["dias_trabajados"]}\nHE: {fila["horas_extra_semana"]}')
        datos.append(renglon)

    col_widths = [0.55 * inch, 1.55 * inch] + [0.82 * inch] * len(dias_semana) + [1.05 * inch]
    tabla = Table(datos, repeatRows=1, colWidths=col_widths)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7.5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F8FAFC'), colors.white]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elementos.append(tabla)

    documento.build(elementos, onFirstPage=_dibujar_marco_pdf, onLaterPages=_dibujar_marco_pdf)
    response.write(buffer.getvalue())
    buffer.close()
    return response


def _contexto_reporte_jornadas(nombre_usuario, filtros, jornadas, catalogos, titulo_pagina, descripcion):
    return {
        'nombre_usuario': nombre_usuario,
        'jornadas': jornadas,
        'filtros': filtros,
        'total_jornadas': jornadas.count(),
        'total_retardos': jornadas.filter(retardo=True).count(),
        'total_con_extra': jornadas.filter(horas_extra__gt=0).count(),
        'titulo_pagina': titulo_pagina,
        'descripcion_reporte': descripcion,
        **catalogos,
    }

def sincronizar_checadas_view(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    if request.method != 'POST':
        return redirect('inicio')

    try:
        resumen = sincronizar_checadas()
        messages.success(
            request,
            (
                f"Sincronizacion completada. Nuevas: {resumen['creados']}, "
                f"omitidas por margen: {resumen['omitidos_margen']}, "
                f"sin empleado: {resumen['sin_empleado']}."
            )
        )
    except Exception as error:
        messages.error(request, f"Error al sincronizar checadas: {error}")

    return redirect('inicio')


def asistencias_hoy(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    nombreUsuarioLogueado = request.session['usuario_nombre']
    filtros, jornadas = _construir_filtros_asistencias(request)
    catalogos = _obtener_catalogos_asistencias()

    context = _contexto_reporte_jornadas(
        nombreUsuarioLogueado,
        filtros,
        jornadas,
        catalogos,
        'Asistencias del dia',
        'Consulta operativa de entradas, salidas, retardos y horas extra por fecha administrativa.',
    )
    return render(request, 'appChecador/asistencias/asistenciasHoy.html', context)


def exportar_asistencias_excel(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros, jornadas = _construir_filtros_asistencias(request)
    return _crear_excel_reporte_jornadas(list(jornadas), filtros, 'Asistencias del dia', 'reporte_asistencias')


def exportar_asistencias_pdf(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros, jornadas = _construir_filtros_asistencias(request)
    return _crear_pdf_reporte_jornadas(list(jornadas), filtros, 'Reporte de Asistencias del Dia', 'reporte_asistencias')


def exportar_asistencias_excel_imss(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros, jornadas = _construir_filtros_asistencias(request)
    return _crear_excel_reporte_jornadas(list(jornadas), filtros, 'Asistencias del dia', 'reporte_asistencias', modo_imss=True)


def exportar_asistencias_pdf_imss(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros, jornadas = _construir_filtros_asistencias(request)
    return _crear_pdf_reporte_jornadas(list(jornadas), filtros, 'Reporte de Asistencias del Dia', 'reporte_asistencias', modo_imss=True)


def retardos(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    nombreUsuarioLogueado = request.session['usuario_nombre']
    filtros, jornadas = _construir_filtros_asistencias(request)
    jornadas = jornadas.filter(retardo=True)
    catalogos = _obtener_catalogos_asistencias()

    context = _contexto_reporte_jornadas(
        nombreUsuarioLogueado,
        filtros,
        jornadas,
        catalogos,
        'Retardos',
        'Empleados con llegada tardia en la fecha administrativa seleccionada.',
    )
    return render(request, 'appChecador/asistencias/retardos.html', context)


def exportar_retardos_excel(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros, jornadas = _construir_filtros_asistencias(request)
    return _crear_excel_reporte_jornadas(list(jornadas.filter(retardo=True)), filtros, 'Retardos', 'reporte_retardos')


def exportar_retardos_pdf(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros, jornadas = _construir_filtros_asistencias(request)
    return _crear_pdf_reporte_jornadas(list(jornadas.filter(retardo=True)), filtros, 'Reporte de Retardos', 'reporte_retardos')


def exportar_retardos_excel_imss(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros, jornadas = _construir_filtros_asistencias(request)
    return _crear_excel_reporte_jornadas(list(jornadas.filter(retardo=True)), filtros, 'Retardos', 'reporte_retardos', modo_imss=True)


def exportar_retardos_pdf_imss(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros, jornadas = _construir_filtros_asistencias(request)
    return _crear_pdf_reporte_jornadas(list(jornadas.filter(retardo=True)), filtros, 'Reporte de Retardos', 'reporte_retardos', modo_imss=True)


def horas_extra(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    nombreUsuarioLogueado = request.session['usuario_nombre']
    filtros, jornadas = _construir_filtros_asistencias(request)
    jornadas = jornadas.filter(horas_extra__gt=0)
    catalogos = _obtener_catalogos_asistencias()

    context = _contexto_reporte_jornadas(
        nombreUsuarioLogueado,
        filtros,
        jornadas,
        catalogos,
        'Horas extra',
        'Empleados con horas extra registradas en la fecha administrativa seleccionada.',
    )
    return render(request, 'appChecador/asistencias/horasExtra.html', context)


def resumen_departamento(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    nombreUsuarioLogueado = request.session['usuario_nombre']
    filtros = _construir_filtros_resumen(request)
    jornadas = JornadaAsistencia.objects.filter(fecha_administrativa=filtros['fecha']).select_related(
        'empleado', 'empleado__departamento', 'horario'
    )
    jornadas = _aplicar_filtros_resumen(jornadas, filtros)
    resumen = _resumen_departamentos(jornadas)

    return render(
        request,
        'appChecador/asistencias/resumenDepartamento.html',
        {
            'nombre_usuario': nombreUsuarioLogueado,
            'filtros': filtros,
            'departamentos': Departamento.objects.order_by('nombre'),
            'empleados': Empleado.objects.order_by('nombre'),
            'resumen': resumen,
            'total_departamentos': len(resumen),
            'total_jornadas': sum(item['total_jornadas'] for item in resumen),
            'total_retardos': sum(item['retardos'] for item in resumen),
        }
    )


def resumen_empleado(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    nombreUsuarioLogueado = request.session['usuario_nombre']
    filtros = _construir_filtros_resumen(request)
    jornadas = JornadaAsistencia.objects.filter(fecha_administrativa=filtros['fecha']).select_related(
        'empleado', 'empleado__departamento', 'horario'
    )
    jornadas = _aplicar_filtros_resumen(jornadas, filtros)
    resumen = _resumen_empleados(jornadas)

    return render(
        request,
        'appChecador/asistencias/resumenEmpleado.html',
        {
            'nombre_usuario': nombreUsuarioLogueado,
            'filtros': filtros,
            'departamentos': Departamento.objects.order_by('nombre'),
            'empleados': Empleado.objects.order_by('nombre'),
            'resumen': resumen,
            'total_empleados': len(resumen),
            'total_retardos': sum(1 for item in resumen if item['retardo']),
            'total_horas_extra': sum(item['horas_extra'] for item in resumen),
        }
    )


def semana_operativa(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    nombreUsuarioLogueado = request.session['usuario_nombre']
    filtros = _construir_filtros_resumen(request)
    inicio_semana, fin_semana = RegistroAsistencia.obtener_semana_operativa(filtros['fecha_base'])
    jornadas = JornadaAsistencia.objects.filter(
        fecha_administrativa__range=(inicio_semana, fin_semana)
    ).select_related('empleado', 'empleado__departamento', 'horario')
    jornadas = _aplicar_filtros_resumen(jornadas, filtros)
    resumen = _resumen_semana_operativa(jornadas)
    dias_semana, filas_semana = _detalle_semana_operativa(jornadas, inicio_semana, fin_semana)

    return render(
        request,
        'appChecador/asistencias/semanaOperativa.html',
        {
            'nombre_usuario': nombreUsuarioLogueado,
            'filtros': filtros,
            'departamentos': Departamento.objects.order_by('nombre'),
            'empleados': Empleado.objects.order_by('nombre'),
            'resumen': resumen,
            'dias_semana': dias_semana,
            'filas_semana': filas_semana,
            'inicio_semana': inicio_semana,
            'fin_semana': fin_semana,
            'total_empleados': len(resumen),
            'total_jornadas': sum(item['total_jornadas'] for item in resumen),
            'total_retardos': sum(item['retardos'] for item in resumen),
            'total_horas_extra': sum(item['horas_extra'] for item in resumen),
        }
    )


def exportar_horas_extra_excel(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros, jornadas = _construir_filtros_asistencias(request)
    return _crear_excel_reporte_jornadas(list(jornadas.filter(horas_extra__gt=0)), filtros, 'Horas extra', 'reporte_horas_extra')


def exportar_horas_extra_pdf(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros, jornadas = _construir_filtros_asistencias(request)
    return _crear_pdf_reporte_jornadas(list(jornadas.filter(horas_extra__gt=0)), filtros, 'Reporte de Horas Extra', 'reporte_horas_extra')


def exportar_horas_extra_excel_imss(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros, jornadas = _construir_filtros_asistencias(request)
    return _crear_excel_reporte_jornadas(list(jornadas.filter(horas_extra__gt=0)), filtros, 'Horas extra', 'reporte_horas_extra', modo_imss=True)


def exportar_horas_extra_pdf_imss(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros, jornadas = _construir_filtros_asistencias(request)
    return _crear_pdf_reporte_jornadas(list(jornadas.filter(horas_extra__gt=0)), filtros, 'Reporte de Horas Extra', 'reporte_horas_extra', modo_imss=True)


def exportar_semana_operativa_excel(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros = _construir_filtros_resumen(request)
    inicio_semana, fin_semana = RegistroAsistencia.obtener_semana_operativa(filtros['fecha_base'])
    jornadas = JornadaAsistencia.objects.filter(
        fecha_administrativa__range=(inicio_semana, fin_semana)
    ).select_related('empleado', 'empleado__departamento', 'horario')
    jornadas = _aplicar_filtros_resumen(jornadas, filtros)
    dias_semana, filas_semana = _detalle_semana_operativa(jornadas, inicio_semana, fin_semana)
    return _crear_excel_semana_operativa_detallada(
        dias_semana,
        filas_semana,
        inicio_semana,
        fin_semana,
        'reporte_semana_operativa',
        filtros,
    )


def exportar_semana_operativa_pdf(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros = _construir_filtros_resumen(request)
    inicio_semana, fin_semana = RegistroAsistencia.obtener_semana_operativa(filtros['fecha_base'])
    jornadas = JornadaAsistencia.objects.filter(
        fecha_administrativa__range=(inicio_semana, fin_semana)
    ).select_related('empleado', 'empleado__departamento', 'horario')
    jornadas = _aplicar_filtros_resumen(jornadas, filtros)
    dias_semana, filas_semana = _detalle_semana_operativa(jornadas, inicio_semana, fin_semana)
    return _crear_pdf_semana_operativa_detallada(
        dias_semana,
        filas_semana,
        inicio_semana,
        fin_semana,
        'reporte_semana_operativa',
        filtros,
    )


def exportar_semana_operativa_excel_imss(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros = _construir_filtros_resumen(request)
    inicio_semana, fin_semana = RegistroAsistencia.obtener_semana_operativa(filtros['fecha_base'])
    jornadas = JornadaAsistencia.objects.filter(
        fecha_administrativa__range=(inicio_semana, fin_semana)
    ).select_related('empleado', 'empleado__departamento', 'horario')
    jornadas = _aplicar_filtros_resumen(jornadas, filtros)
    dias_semana, filas_semana = _detalle_semana_operativa(jornadas, inicio_semana, fin_semana, modo_imss=True)
    return _crear_excel_semana_operativa_detallada(
        dias_semana,
        filas_semana,
        inicio_semana,
        fin_semana,
        'reporte_semana_operativa',
        filtros,
        modo_imss=True,
    )


def exportar_semana_operativa_pdf_imss(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    filtros = _construir_filtros_resumen(request)
    inicio_semana, fin_semana = RegistroAsistencia.obtener_semana_operativa(filtros['fecha_base'])
    jornadas = JornadaAsistencia.objects.filter(
        fecha_administrativa__range=(inicio_semana, fin_semana)
    ).select_related('empleado', 'empleado__departamento', 'horario')
    jornadas = _aplicar_filtros_resumen(jornadas, filtros)
    dias_semana, filas_semana = _detalle_semana_operativa(jornadas, inicio_semana, fin_semana, modo_imss=True)
    return _crear_pdf_semana_operativa_detallada(
        dias_semana,
        filas_semana,
        inicio_semana,
        fin_semana,
        'reporte_semana_operativa',
        filtros,
        modo_imss=True,
    )

def usuarios(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    nombreUsuarioLogueado = request.session['usuario_nombre']
    error_creacion = None

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        is_superuser = request.POST.get('is_superuser') == 'on'

        if not username or not password:
            error_creacion = 'El usuario y la contrasena son obligatorios.'
            messages.error(request, error_creacion)
        elif password != password_confirm:
            error_creacion = 'Las contrasenas no coinciden.'
            messages.error(request, error_creacion)
        elif User.objects.filter(username=username).exists():
            error_creacion = 'Ese nombre de usuario ya existe.'
            messages.error(request, error_creacion)
        else:
            User.objects.create_user(
                username=username,
                password=password,
                first_name=first_name,
                last_name=last_name,
                email=email,
                is_staff=is_superuser,
                is_superuser=is_superuser,
            )
            messages.success(request, 'Usuario registrado')
            return redirect('usuarios')

    usuarios = User.objects.all().order_by('username')

    return render(
        request,
        'appChecador/usuarios/usuarios.html',
        {
            'nombre_usuario': nombreUsuarioLogueado,
            'usuarios': usuarios,
            'error_creacion': error_creacion,
        }
    )

def empleados(request):
    if 'usuario_id' not in request.session:
        return redirect('login')
    
    nombreUsuarioLogueado = request.session['usuario_nombre']
    
    #Cargar empleados desde la base de datos
    empleados = Empleado.objects.select_related('departamento', 'horario').all()
    
    return render(request, 'appChecador/checador/empleados.html', {'nombre_usuario': nombreUsuarioLogueado, 'empleados': empleados})

def importar_empleados_sin_configurar(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    if request.method != 'POST':
        return redirect('empleados')

    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        messages.error(request, 'Selecciona un archivo Excel para importar.')
        return redirect('empleados')

    try:
        workbook = load_workbook(archivo)
        hoja = workbook.active
    except Exception:
        messages.error(request, 'No se pudo leer el archivo Excel.')
        return redirect('empleados')

    actualizados = 0
    omitidos_no_encontrados = 0
    omitidos_sin_id = 0

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        id_biometrico = str(fila[0]).strip() if fila[0] is not None else ''
        if not id_biometrico:
            omitidos_sin_id += 1
            continue

        empleado = Empleado.objects.filter(id_biometrico=id_biometrico).first()
        if not empleado:
            omitidos_no_encontrados += 1
            continue

        empleado.puesto = str(fila[2]).strip() if len(fila) > 2 and fila[2] is not None else None
        empleado.numero_empleado = str(fila[3]).strip() if len(fila) > 3 and fila[3] is not None else None
        empleado.numero_seguridad_social = str(fila[4]).strip() if len(fila) > 4 and fila[4] is not None else None
        empleado.save()
        actualizados += 1

    messages.success(
        request,
        (
            f"Importacion completada. Actualizados: {actualizados}, "
            f"omitidos por no existir: {omitidos_no_encontrados}, "
            f"sin ID: {omitidos_sin_id}."
        )
    )
    return redirect('empleados')

def exportar_empleados_sin_configurar(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    empleados = (
        Empleado.objects.filter(Q(puesto__isnull=True) | Q(puesto=''))
        .filter(Q(numero_empleado__isnull=True) | Q(numero_empleado=''))
        .filter(Q(numero_seguridad_social__isnull=True) | Q(numero_seguridad_social=''))
        .order_by('nombre', 'id_biometrico')
    )

    workbook = Workbook()
    hoja = workbook.active
    hoja.title = "Empleados sin configurar"
    hoja.append([
        'ID Biometrico',
        'Nombre',
        'Puesto',
        'Numero Empleado',
        'Numero Seguridad Social',
    ])

    for empleado in empleados:
        hoja.append([
            empleado.id_biometrico,
            empleado.nombre,
            empleado.puesto or '',
            empleado.numero_empleado or '',
            empleado.numero_seguridad_social or '',
        ])

    for columna in hoja.columns:
        max_len = 0
        letra_columna = columna[0].column_letter
        for celda in columna:
            valor = str(celda.value or '')
            if len(valor) > max_len:
                max_len = len(valor)
        hoja.column_dimensions[letra_columna].width = max_len + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=\"empleados_sin_configurar.xlsx\"'
    workbook.save(response)
    return response

def departamentos(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    nombreUsuarioLogueado = request.session['usuario_nombre']

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        usa_corte_madrugada = request.POST.get('usa_corte_madrugada') == 'on'

        if nombre:
            Departamento.objects.create(
                nombre=nombre,
                usa_corte_madrugada=usa_corte_madrugada,
            )
            messages.success(request, 'Departamento registrado')
        else:
            messages.error(request, 'Error al registrar el departamento')
        return redirect('departamentos')

    departamentos = Departamento.objects.annotate(
        total_empleados=Count('empleado')
    ).order_by('nombre')

    return render(
        request,
        'appChecador/checador/departamentos.html',
        {
            'nombre_usuario': nombreUsuarioLogueado,
            'departamentos': departamentos,
        }
    )

def horarios(request):
    if 'usuario_id' not in request.session:
        return redirect('login')

    nombreUsuarioLogueado = request.session['usuario_nombre']

    if request.method == 'POST':
        nombre_turno = request.POST.get('nombre_turno', '').strip()
        dia_semana = request.POST.get('dia_semana') or None
        hora_entrada = request.POST.get('hora_entrada')
        hora_salida = request.POST.get('hora_salida')
        usa_corte_madrugada = request.POST.get('usa_corte_madrugada') == 'on'
        tolerancia_entrada = request.POST.get('tolerancia_entrada') or 15
        inicio_entrada = request.POST.get('inicio_entrada') or None
        minimo_minutos_para_contar_extra = request.POST.get('minimo_minutos_para_contar_extra') or 59

        if nombre_turno and hora_entrada and hora_salida:
            Horario.objects.create(
                nombre_turno=nombre_turno,
                dia_semana=dia_semana,
                hora_entrada=hora_entrada,
                hora_salida=hora_salida,
                usa_corte_madrugada=usa_corte_madrugada,
                tolerancia_entrada=tolerancia_entrada,
                inicio_entrada=inicio_entrada,
                minimo_minutos_para_contar_extra=minimo_minutos_para_contar_extra,
            )
            messages.success(request, 'Horario registrado')
        else:
            messages.error(request, 'Error al registrar el horario')
        return redirect('horarios')

    horarios = Horario.objects.annotate(
        total_empleados=Count('empleados_asignados', distinct=True)
    ).order_by('dia_semana', 'hora_entrada', 'nombre_turno')

    return render(
        request,
        'appChecador/checador/horarios.html',
        {
            'nombre_usuario': nombreUsuarioLogueado,
            'horarios': horarios,
            'dias_semana': Horario.DIAS_SEMANA,
        }
    )


def editar_horario(request, horario_id):
    if 'usuario_id' not in request.session:
        return redirect('login')

    nombreUsuarioLogueado = request.session['usuario_nombre']
    horario = get_object_or_404(Horario, id=horario_id)

    if request.method == 'POST':
        horario.nombre_turno = request.POST.get('nombre_turno', '').strip()
        horario.dia_semana = request.POST.get('dia_semana') or None
        horario.hora_entrada = request.POST.get('hora_entrada')
        horario.hora_salida = request.POST.get('hora_salida')
        horario.usa_corte_madrugada = request.POST.get('usa_corte_madrugada') == 'on'
        horario.tolerancia_entrada = request.POST.get('tolerancia_entrada') or 15
        horario.inicio_entrada = request.POST.get('inicio_entrada') or None
        horario.minimo_minutos_para_contar_extra = request.POST.get('minimo_minutos_para_contar_extra') or 59
        horario.save()
        messages.success(request, 'Horario actualizado')
        return redirect('editar_horario', horario_id=horario.id)

    return render(
        request,
        'appChecador/horarios/editarHorario.html',
        {
            'nombre_usuario': nombreUsuarioLogueado,
            'horario': horario,
            'dias_semana': Horario.DIAS_SEMANA,
        }
    )

def cambiar_password_usuario(request, usuario_id):
    if 'usuario_id' not in request.session:
        return redirect('login')

    nombreUsuarioLogueado = request.session['usuario_nombre']
    usuario = get_object_or_404(User, id=usuario_id)
    error_password = None

    if request.method == 'POST':
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')

        if not password:
            error_password = 'La contrasena no puede estar vacia.'
        elif password != password_confirm:
            error_password = 'Las contrasenas no coinciden.'
        else:
            usuario.set_password(password)
            usuario.save()
            return redirect('usuarios')

    return render(
        request,
        'appChecador/usuarios/cambiarPasswordUsuario.html',
        {
            'nombre_usuario': nombreUsuarioLogueado,
            'usuario_sistema': usuario,
            'error_password': error_password,
        }
    )

def editar_empleado(request, empleado_id):
    if 'usuario_id' not in request.session:
        return redirect('login')

    nombreUsuarioLogueado = request.session['usuario_nombre']
    empleado = get_object_or_404(Empleado, id=empleado_id)
    departamentos = Departamento.objects.all().order_by('nombre')
    horarios_por_departamento = _agrupar_horarios_por_departamento()

    if request.method == 'POST':
        empleado.nombre = request.POST.get('nombre', '').strip()
        empleado.puesto = request.POST.get('puesto', '').strip() or None
        empleado.numero_empleado = request.POST.get('numero_empleado', '').strip() or None
        empleado.numero_seguridad_social = request.POST.get('numero_seguridad_social', '').strip() or None
        empleado.activo = request.POST.get('activo') == 'on'

        departamento_id = request.POST.get('departamento')
        if departamento_id:
            empleado.departamento = Departamento.objects.filter(id=departamento_id).first()
        else:
            empleado.departamento = None

        empleado.save()

        horario_id = request.POST.get('horario')
        empleado.horario = Horario.objects.filter(id=horario_id).first() if horario_id else None
        empleado.save()
        reprocesar_jornadas_empleado(empleado)
        messages.success(request, 'Empleado actualizado y asistencias recalculadas')

        return redirect('editar_empleado', empleado_id=empleado.id)

    return render(
        request,
        'appChecador/empleados/editarEmpleado.html',
        {
            'nombre_usuario': nombreUsuarioLogueado,
            'empleado': empleado,
            'departamentos': departamentos,
            'horarios_por_departamento': horarios_por_departamento,
        }
    )


    
